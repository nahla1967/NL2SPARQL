"""
eval_metrics.py
----------------
Reads eval_results.jsonl (produced by eval_runner.py) and writes
NL2SPARQL_Evaluation_Results.xlsx with two sheets:

    Raw     — one row per run, exactly as logged.
    Summary — Execution Accuracy / SPARQL Validity / Exact Match / F1,
              broken down by language, by strategy (single_kg1/2 only,
              where strategy is meaningful), by tier, and by category.
              All Summary numbers are COUNTIFS/AVERAGEIFS formulas
              pointing at the Raw sheet, so re-running eval_runner.py and
              re-pasting the Raw data recalculates Summary automatically —
              nothing here is a hardcoded Python-computed number.

Run this after eval_runner.py. Works fine on partial results too (useful
while the dataset is still being filled in incrementally).
"""

import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Anchored to this script's own folder (not the current working directory)
# so RESULTS_PATH/OUT_PATH resolve to the same file regardless of which
# directory you launch this script from — matches eval_runner.py.
_HERE = os.path.dirname(os.path.abspath(__file__))
RESULTS_PATH = os.path.join(_HERE, "results", "eval_results.jsonl")
OUT_PATH     = os.path.join(_HERE, "results", "NL2SPARQL_Evaluation_Results.xlsx")

FONT_NAME   = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="2D3B4F")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
BODY_FONT   = Font(name=FONT_NAME, size=10)
SECTION_FONT= Font(name=FONT_NAME, size=11, bold=True, color="2D3B4F")
THIN        = Side(style="thin", color="C9CCD1")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT_FORMAT  = "0.0%"

RAW_COLUMNS = [
    "id", "tier", "category", "kg", "language", "strategy",
    "expected_type", "query_type", "routing_ok", "sparql_valid",
    "failure_type", "error_detail", "exact_match", "f1", "duration_s",
]

LANGUAGES  = ["en", "fr", "ar"]
STRATEGIES = ["zero-shot", "few-shot", "cot"]
CATEGORIES = [
    "single_kg1", "single_kg2", "single_kg3", "cross_kg", "open_kg",
    "count_kg1", "count_kg2", "count_kg3",
    "filter_numeric_kg1", "filter_numeric_kg2", "filter_numeric_kg3",
    "filter_string_kg2", "filter_string_kg3",
    "ranking_kg1", "ranking_kg2", "ranking_kg3",
    "compare_two_airports", "compare_two_flights", "compare_two_departments",
    "group_aggregate_kg1", "group_aggregate_kg2", "group_aggregate_kg3",
    "out_of_scope", "ask_query", "typo_fuzzy",
    "multilingual_edge", "property_ambiguity",
    "ghost_property", "multi_value_ambiguity",
]

def load_records():
    records = []
    with open(RESULTS_PATH, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                records.append(json.loads(line))
    return records


def write_raw_sheet(wb, records):
    ws = wb.create_sheet("Raw")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    for c, name in enumerate(RAW_COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = 16

    for r, rec in enumerate(records, start=2):
        for c, name in enumerate(RAW_COLUMNS, start=1):
            val = rec.get(name)
            if isinstance(val, (dict, list)):
                val = json.dumps(val, ensure_ascii=False)
            if name == "exact_match" and isinstance(val, bool):
                # Excel's AVERAGEIFS ignores boolean TRUE/FALSE cells in the
                # averaged range (unlike COUNTIFS, which matches them fine).
                # exact_match() returns bool, so it must be cast to 1/0 here
                # or the Summary sheet's "Exact Match" column stays blank.
                val = int(val)
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER

    ws.auto_filter.ref = f"A1:{get_column_letter(len(RAW_COLUMNS))}{max(len(records) + 1, 1)}"
    return len(records)


def _col(name):
    return RAW_COLUMNS.index(name) + 1  # 1-indexed


def add_summary_block(ws, start_row, title, group_col, group_values, n_raw_rows):
    """
    Writes one summary block (e.g. 'By language') as a small table with
    live formulas referencing the Raw sheet. Returns the next free row.
    """
    last_raw = n_raw_rows + 1  # +1 for header row

    routing_col   = get_column_letter(_col("routing_ok"))
    sparql_col    = get_column_letter(_col("sparql_valid"))
    failure_col   = get_column_letter(_col("failure_type"))
    em_col        = get_column_letter(_col("exact_match"))
    f1_col        = get_column_letter(_col("f1"))
    group_col_ltr = get_column_letter(_col(group_col))

    ws.cell(row=start_row, column=1, value=title).font = SECTION_FONT
    r = start_row + 1

    headers = ["Group", "N runs", "Routing Accuracy", "Execution Accuracy",
               "SPARQL Validity", "Exact Match", "F1 (avg)"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
    r += 1

    for gv in group_values:
        rng = lambda col: f"Raw!{col}2:{col}{last_raw}"
        grp_rng = rng(group_col_ltr)

        n_formula        = f'=COUNTIF({grp_rng},"{gv}")'
        routing_formula  = (f'=IFERROR(COUNTIFS({grp_rng},"{gv}",{rng(routing_col)},TRUE)'
                             f'/COUNTIF({grp_rng},"{gv}"),"")')
        exec_formula     = (f'=IFERROR(COUNTIFS({grp_rng},"{gv}",{rng(failure_col)},"success")'
                             f'/COUNTIF({grp_rng},"{gv}"),"")')
        sparql_formula   = (f'=IFERROR(COUNTIFS({grp_rng},"{gv}",{rng(sparql_col)},TRUE)'
                             f'/COUNTIF({grp_rng},"{gv}"),"")')
        em_formula       = f'=IFERROR(AVERAGEIFS({rng(em_col)},{grp_rng},"{gv}"),"")'
        f1_formula       = f'=IFERROR(AVERAGEIFS({rng(f1_col)},{grp_rng},"{gv}"),"")'

        row_values = [gv, n_formula, routing_formula, exec_formula,
                      sparql_formula, em_formula, f1_formula]
        for c, val in enumerate(row_values, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
            if c >= 3:
                cell.number_format = PCT_FORMAT if c != 2 else "0"
                cell.alignment = Alignment(horizontal="center")
        r += 1

    return r + 1  # blank row after the block


def write_summary_sheet(wb, n_raw_rows):
    ws = wb.create_sheet("Summary", 0)  # make it the first visible sheet
    ws.sheet_view.showGridLines = False

    ws["A1"] = "NL2SPARQL — Evaluation Results Summary"
    ws["A1"].font = Font(name=FONT_NAME, size=14, bold=True)
    ws["A2"] = ("All figures are live formulas over the Raw sheet — re-running eval_runner.py "
                "and pasting fresh Raw data recalculates everything below.")
    ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")

    for c, w in enumerate([26, 10, 18, 18, 16, 14, 12], start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    row = 4
    row = add_summary_block(ws, row, "Overall", "language", ["en", "fr", "ar"], n_raw_rows)
    row = add_summary_block(ws, row, "By strategy (single_kg1 / single_kg2 only — "
                                      "other branches don't vary by strategy)",
                             "strategy", STRATEGIES, n_raw_rows)
    row = add_summary_block(ws, row, "By tier", "tier", [1, 2, 3], n_raw_rows)
    row = add_summary_block(ws, row, "By category", "category", CATEGORIES, n_raw_rows)


def main():
    records = load_records()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop default empty sheet

    n_raw = write_raw_sheet(wb, records)
    write_summary_sheet(wb, n_raw)

    wb.save(OUT_PATH)
    print(f"Saved {OUT_PATH} — {n_raw} raw runs, Summary sheet has live formulas.")
    print("Run scripts/recalc.py on this file (LibreOffice) to populate cached values "
          "before opening it read-only elsewhere, e.g.:")
    print(f"  python3 /mnt/skills/public/xlsx/scripts/recalc.py {OUT_PATH}")


if __name__ == "__main__":
    main()