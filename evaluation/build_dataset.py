"""
build_dataset.py
-----------------
Generates NL2SPARQL_Evaluation_Dataset.xlsx — the fillable template for the
72-canonical-question evaluation set (Tier 1 / 2 / 3, per the agreed plan).

Run once to produce the empty (mostly) template. Esc fills in the yellow
cells with real questions + ground-truth answers, then eval_runner.py
consumes this file.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT_PATH = "/home/claude/eval/NL2SPARQL_Evaluation_Dataset.xlsx"

# ── STYLE CONSTANTS ────────────────────────────────────────────────────────
FONT_NAME   = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="2D3B4F")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
EDIT_FILL   = PatternFill("solid", fgColor="FFF6C9")   # yellow = fill this in
LOCK_FILL   = PatternFill("solid", fgColor="F2F2F2")   # grey = pre-filled, don't edit
EXAMPLE_FILL= PatternFill("solid", fgColor="DCEEDC")   # green = worked example row
BODY_FONT   = Font(name=FONT_NAME, size=10)
THIN        = Side(style="thin", color="C9CCD1")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("id",                  10),
    ("tier",                6),
    ("category",            22),
    ("kg",                  12),
    ("expected_type",       16),
    ("template_name",       20),
    ("strategy_applicable", 12),
    ("question_en",         46),
    ("question_fr",         46),
    ("question_ar",         46),
    ("expected_answer",     22),
    ("notes",               30),
]
EDITABLE_COLS = {"question_en", "question_fr", "question_ar", "expected_answer", "notes"}

# ── TIER / CATEGORY DEFINITIONS ────────────────────────────────────────────
# (category, kg, expected_type, template_name, strategy_applicable, count, tier)
SPEC = [
    ("single_kg1",           "flights",    "single_kg1", None,                   True,  9,  1),
    ("single_kg2",           "airports",   "single_kg2", None,                   True,  9,  1),
    ("cross_kg",             "cross",      "cross_kg",   None,                   False, 12, 1),
    ("open_kg",              "cross",      "open_kg",    None,                   False, 6,  1),
    ("count_kg1",            "flights",    "template",   "count_kg1",            False, 3,  2),
    ("filter_numeric_kg1",   "flights",    "template",   "filter_numeric_kg1",   False, 3,  2),
    ("filter_numeric_kg2",   "airports",   "template",   "filter_numeric_kg2",   False, 3,  2),
    ("filter_string_kg2",    "airports",   "template",   "filter_string_kg2",    False, 3,  2),
    ("ranking_kg2",          "airports",   "template",   "ranking_kg2",          False, 3,  2),
    ("compare_two_airports", "airports",   "template",   "compare_two_airports", False, 3,  2),
    ("out_of_scope",         "n/a",        "out_of_scope", None,                 False, 3,  3),
    ("ask_query",            "mixed",      "ask_query",  None,                   False, 4,  3),
    ("typo_fuzzy",           "mixed",      "VARIES",     None,                   False, 3,  3),
    ("multilingual_edge",    "mixed",      "VARIES",     None,                   False, 4,  3),
    ("property_ambiguity",   "mixed",      "VARIES",     None,                   False, 4,  3),
]

# Worked examples (real data, drawn from the existing KG1/KG2/KG3 designs)
# so the format is unambiguous. Everything else is a placeholder for Esc
# to fill with real values from the live Fuseki datasets.
EXAMPLES = {
    "single_kg1_001": dict(
        question_en="What is the departure city of flight OS295?",
        question_fr="Quelle est la ville de départ du vol OS295?",
        question_ar="ما هي مدينة مغادرة الرحلة OS295؟",
        expected_answer="Vienna",
        notes="Worked example — replace all other rows in this block with real flight numbers.",
    ),
    "single_kg2_001": dict(
        question_en="What is the elevation of VIE?",
        question_fr="Quelle est l'élévation de VIE?",
        question_ar="ما هو ارتفاع مطار VIE؟",
        expected_answer="600",
        notes="Worked example — value must match the real elevationFt in KG2 for VIE.",
    ),
    "cross_kg_001": dict(
        question_en="What country does flight OS295 land in?",
        question_fr="Dans quel pays atterrit le vol OS295?",
        question_ar="في أي دولة يهبط الرحلة OS295؟",
        expected_answer="Austria",
        notes="Worked example — cross-KG bridge via destination IATA.",
    ),
}


def build():
    wb = openpyxl.Workbook()

    # ── SHEET 1: Legend ────────────────────────────────────────────────────
    legend = wb.active
    legend.title = "Legend"
    legend.sheet_view.showGridLines = False

    legend["A1"] = "NL2SPARQL Evaluation Dataset — Legend"
    legend["A1"].font = Font(name=FONT_NAME, size=14, bold=True)
    legend.merge_cells("A1:D1")

    legend_rows = [
        ("", ""),
        ("Color key", ""),
        ("Yellow cells", "Fill in with real data (question text per language, ground-truth answer, notes)."),
        ("Grey cells", "Pre-filled metadata — do not edit (id, tier, category, routing expectations)."),
        ("Green rows", "Worked examples showing the expected format — do not delete, just reference."),
        ("", ""),
        ("Column reference", ""),
        ("id", "Unique row identifier, e.g. single_kg1_001."),
        ("tier", "1 = core claims, 2 = template patterns, 3 = robustness/edge cases."),
        ("category", "Sub-branch being tested — matches router.py / TEMPLATE_REGISTRY names."),
        ("kg", "Which knowledge graph(s) the question targets (flights / airports / university / cross / mixed / n/a)."),
        ("expected_type", "The query_type the router SHOULD return. 'VARIES' means it depends on which "
                           "branch the specific edge case is stress-testing — set per-row when you fill it in."),
        ("template_name", "Only set for Tier 2 template rows — matches TEMPLATE_REGISTRY keys."),
        ("strategy_applicable", "TRUE only for single_kg1 / single_kg2 (and single_kg3 if added) — these are "
                                 "the only branches where zero-shot / few-shot / cot actually change the SPARQL. "
                                 "Everything else runs once per language, not three times."),
        ("question_en / _fr / _ar", "The real question text in each language. Must reference entities that "
                                     "actually exist in your live KGs (real flight numbers, real IATA codes, "
                                     "real LUBM entity names) — the whole point is these must resolve."),
        ("expected_answer", "The ground-truth value (not full sentence) used for Exact Match / F1 scoring — "
                             "e.g. 'Vienna', '1738', 'Yes', a count number. Keep it language-neutral where "
                             "possible so scoring is consistent across en/fr/ar runs."),
        ("notes", "Anything eval_runner.py or a human reviewer should know about this row."),
        ("", ""),
        ("Tier structure (72 canonical questions)", ""),
        ("Tier 1 — Core claims (36)", "single_kg1: 9 · single_kg2: 9 · cross_kg: 12 · open_kg: 6"),
        ("Tier 2 — Template patterns (18)", "count_kg1: 3 · filter_numeric_kg1: 3 · filter_numeric_kg2: 3 · "
                                             "filter_string_kg2: 3 · ranking_kg2: 3 · compare_two_airports: 3"),
        ("Tier 3 — Robustness / edge cases (18)", "out_of_scope: 3 · ask_query: 4 · typo_fuzzy: 3 · "
                                                   "multilingual_edge: 4 · property_ambiguity: 4"),
        ("", ""),
        ("Run expansion (handled by eval_runner.py)", ""),
        ("Per row", "× 3 languages always. × 3 strategies only if strategy_applicable = TRUE."),
        ("Total runs", "single_kg1 (81) + single_kg2 (81) + cross_kg (36) + open_kg (18) + "
                        "Tier 2 (54) + Tier 3 (54) = 324"),
    ]
    r = 3
    for label, desc in legend_rows:
        legend.cell(row=r, column=1, value=label).font = Font(name=FONT_NAME, size=10, bold=bool(desc == ""))
        legend.cell(row=r, column=2, value=desc).font = Font(name=FONT_NAME, size=10)
        legend.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    legend.column_dimensions["A"].width = 34
    legend.column_dimensions["B"].width = 90

    # ── SHEET 2: Questions ─────────────────────────────────────────────────
    ws = wb.create_sheet("Questions")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    for c, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 20

    # Data validation dropdowns
    dv_bool = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
    ws.add_data_validation(dv_bool)

    row_idx = 2
    for category, kg, expected_type, template_name, strategy_applicable, count, tier in SPEC:
        for n in range(1, count + 1):
            rid = f"{category}_{n:03d}"
            is_example = rid in EXAMPLES

            values = {
                "id": rid,
                "tier": tier,
                "category": category,
                "kg": kg,
                "expected_type": expected_type,
                "template_name": template_name or "",
                "strategy_applicable": "TRUE" if strategy_applicable else "FALSE",
                "question_en": "",
                "question_fr": "",
                "question_ar": "",
                "expected_answer": "",
                "notes": "",
            }
            if is_example:
                values.update(EXAMPLES[rid])

            for c, (name, _) in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=row_idx, column=c, value=values[name])
                cell.font = BODY_FONT
                cell.border = BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=(name in EDITABLE_COLS))
                if name in EDITABLE_COLS:
                    cell.fill = EXAMPLE_FILL if is_example else EDIT_FILL
                else:
                    cell.fill = EXAMPLE_FILL if is_example else LOCK_FILL

            dv_bool.add(ws.cell(row=row_idx, column=7))  # strategy_applicable column (locked, informational)
            row_idx += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{row_idx - 1}"

    wb.save(OUT_PATH)
    print(f"Saved {OUT_PATH} with {row_idx - 2} question rows.")


if __name__ == "__main__":
    build()
